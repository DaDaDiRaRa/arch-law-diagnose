"""사전 사업성 검토 결과 → 1장 요약 Markdown / xlsx.

run_feasibility() 응답(result)을 시니어 보고용 한 장 요약으로 변환.
LLM 호출 없음 — 기존 데이터를 표로 렌더링만.
"""
from __future__ import annotations

import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)


def _num(v: Any, d: int = 0) -> str:
    if v is None or v == "":
        return "—"
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    if d == 0:
        return f"{f:,.0f}"
    return f"{f:,.{d}f}"


def _site_source_label(src: str | None) -> str:
    return {
        "user_override": "수동 입력",
        "auto": "자동 조회",
        "default_1000": "기본값(1000㎡)",
    }.get(src or "", src or "—")


def _gap_label(status: str | None) -> str:
    return {
        "ok": "충족",
        "over": "초과",
        "unknown": "확인 불가",
        "no_target": "요구 없음",
    }.get(status or "", status or "—")


# ── Markdown ────────────────────────────────────────────────────────────
def to_markdown(
    result: dict,
    form_data: dict | None = None,
    project_name: str = "",
    company: str = "",
    author: str = "",
) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    address = result.get("address") or "—"
    land = result.get("land_facts") or {}
    proposal = result.get("proposal") or {}
    categories = result.get("categories") or []
    review = result.get("review_burden") or {}
    rec = result.get("overall_recommendation") or {}

    L: list[str] = []
    L.append("# 사전 사업성 검토 요약")
    L.append("")
    if project_name:
        L.append(f"- **프로젝트**: {project_name}")
    L.append(f"- **주소**: {address}")
    L.append(f"- **작성일**: {today}")
    if company or author:
        L.append(f"- **작성**: {company} {author}".rstrip())
    L.append("")

    # 종합 판단
    L.append("## 종합 판단")
    L.append(f"**{rec.get('verdict', '—')}** — {rec.get('reason', '')}")
    L.append("")

    # 대지 정보
    L.append("## 대지 정보")
    L.append("| 항목 | 값 |")
    L.append("| --- | --- |")
    L.append(f"| 용도지역 | {land.get('zone_use') or '미확인'} |")
    L.append(f"| 지역지구 | {land.get('zone_district') or '—'} |")
    L.append(
        f"| 대지면적 | {_num(result.get('site_area_used'))}㎡ "
        f"({_site_source_label(result.get('site_area_source'))}) |"
    )
    L.append("")

    # 가능 범위 제안
    L.append("## 이 대지에 지을 수 있는 범위")
    L.append("| 항목 | 값 | 비고 |")
    L.append("| --- | --- | --- |")
    cov = proposal.get("max_building_coverage_pct")
    L.append(
        f"| 최대 건폐율 | {_num(cov, 1)}% | 최대 건축면적 "
        f"{_num(proposal.get('max_building_area_sqm'))}㎡ |"
    )
    far = proposal.get("far_pct")
    far_relief = proposal.get("max_far_pct_relief")
    far_note = (
        f"완화 시 최대 {_num(far_relief, 1)}%"
        if far_relief and far and far_relief > far else "—"
    )
    L.append(f"| 최대 용적률 | {_num(far, 1)}% | {far_note} |")
    fa = proposal.get("max_floor_area_sqm")
    fa_relief = proposal.get("max_floor_area_relief_sqm")
    fa_note = (
        f"완화 시 {_num(fa_relief)}㎡"
        if fa_relief and fa and fa_relief > fa else "용적률 한도 기준"
    )
    L.append(f"| 가능 연면적 | {_num(fa)}㎡ | {fa_note} |")
    L.append(
        f"| 권장 주차대수 | {_num(proposal.get('recommended_parking_spaces'))}대 | "
        f"최대 연면적 기준 |"
    )
    L.append("")
    relief_items = proposal.get("applied_relief_items") or []
    if relief_items:
        labels = ", ".join(it.get("label") or it.get("kind", "") for it in relief_items)
        L.append(f"적용 완화: {labels}")
        L.append("")

    # 자동 제안 대안 + 산정 근거
    alts = result.get("auto_alternatives") or []
    if alts:
        L.append("## 자동 제안 대안 (연면적 큰 순)")
        L.append("| 대안 | 건폐율 | 용적률 | 가능 연면적 | 권장 주차 | 심의 |")
        L.append("| --- | --- | --- | --- | --- | --- |")
        for a in alts:
            L.append(
                f"| {a.get('label', '')} | {_num(a.get('building_coverage_pct'), 1)}% | "
                f"{_num(a.get('far_pct'), 1)}% | {_num(a.get('max_floor_area_sqm'))}㎡ | "
                f"{_num(a.get('recommended_parking_spaces'))}대 | "
                f"{a.get('review_count_required', 0)}건 |"
            )
        L.append("")
        for a in alts:
            d = a.get("derivation") or {}
            L.append(f"### {a.get('label', '')} — 산정 근거")
            L.append(f"- 기본 용적률 {_num(d.get('base_far_pct'), 1)}%")
            for b in d.get("relief_breakdown") or []:
                basis = f" · {b.get('basis')}" if b.get("basis") else ""
                L.append(
                    f"  - +{_num(b.get('relief_pct'), 1)}% {b.get('label', '')}{basis}"
                )
            if d.get("cap_note"):
                L.append(f"  - 캡 적용: {d.get('cap_note')}")
            L.append(f"- 최종 용적률 {_num(d.get('final_far_pct'), 1)}%")
            if d.get("site_area_sqm"):
                L.append(
                    f"- 연면적 = {_num(d.get('site_area_sqm'))}㎡ × "
                    f"{_num(d.get('final_far_pct'), 1)}% = {_num(a.get('max_floor_area_sqm'))}㎡"
                )
            reqs = a.get("review_required") or []
            if reqs:
                parts = []
                for r in reqs:
                    ref = f" 「{r.get('law_ref')}」" if r.get("law_ref") else ""
                    parts.append(f"{r.get('name', '')}{ref}")
                L.append(f"- 심의·평가 필요: {'; '.join(parts)}")
            L.append("")

    # 갭 분석 (target 있을 때만)
    has_target = any(
        (c.get("gap_analysis") or {}).get("has_target") for c in categories
    )
    if has_target:
        L.append("## 공모 요구치 대비 (갭 분석)")
        L.append("| 카테고리 | 공모 요구 | 법적 가능 | 판정 |")
        L.append("| --- | --- | --- | --- |")
        for c in categories:
            gap = c.get("gap_analysis") or {}
            if not gap.get("has_target"):
                continue
            unit = c.get("unit", "")
            L.append(
                f"| {c.get('label')} | {_num(c.get('competition_target'), 1)}{unit} | "
                f"{_num(c.get('legal_limit'), 1)}{unit} | "
                f"{_gap_label(gap.get('status'))} ({gap.get('gap_text', '')}) |"
            )
        L.append("")

        # 완화 시나리오
        scen_cat = next(
            (c for c in categories if c.get("key") == "far" and c.get("scenarios")),
            None,
        )
        if scen_cat:
            L.append("## 용적률 완화 시나리오")
            L.append("| 완화 항목 | 적용 후 | 충족 |")
            L.append("| --- | --- | --- |")
            for s in scen_cat.get("scenarios", []):
                L.append(
                    f"| {s.get('label', '')} | {_num(s.get('result_pct'), 1)}% | "
                    f"{'충족' if s.get('covers_target') else '부족'} |"
                )
            L.append("")

    # 심의 부담
    L.append("## 심의·평가 부담")
    req_items = review.get("required") or []
    maybe_items = review.get("maybe") or []
    if req_items:
        L.append(f"**필수 ({len(req_items)})**")
        for r in req_items:
            L.append(f"- {r.get('name', '')} — {r.get('reason', '')}")
    if maybe_items:
        L.append("")
        L.append(f"**조건부 ({len(maybe_items)})**")
        for r in maybe_items:
            L.append(f"- {r.get('name', '')} — {r.get('reason', '')}")
    if not req_items and not maybe_items:
        L.append("자동 트리거된 심의·평가 없음.")
    L.append("")

    L.append("---")
    L.append(
        "사전 사업성 검토는 참여 판단 보조용입니다. "
        "실제 인허가 가능성은 시니어 건축사 검토가 필수입니다."
    )
    return "\n".join(L)


# ── xlsx ────────────────────────────────────────────────────────────────
def to_xlsx(
    result: dict,
    form_data: dict | None = None,
    project_name: str = "",
    company: str = "",
    author: str = "",
) -> bytes:
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    today = datetime.now().strftime("%Y-%m-%d")
    land = result.get("land_facts") or {}
    proposal = result.get("proposal") or {}
    categories = result.get("categories") or []
    review = result.get("review_burden") or {}
    rec = result.get("overall_recommendation") or {}

    header_fill = PatternFill(start_color="F1F3F5", end_color="F1F3F5", fill_type="solid")
    accent_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=16)
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # ── Sheet 1: 사업성 요약 ────────────────────────────────────────────
    ws = wb.active
    ws.title = "사업성요약"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 30

    r = 1
    ws.cell(r, 1, "사전 사업성 검토 요약").font = title_font
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1
    for label, val in [
        ("프로젝트", project_name or "—"),
        ("주소", result.get("address") or "—"),
        ("작성일", today),
        ("작성", f"{company} {author}".strip() or "—"),
    ]:
        ws.cell(r, 1, label).font = bold
        ws.cell(r, 2, val)
        r += 1
    r += 1

    def section(title: str):
        nonlocal r
        c = ws.cell(r, 1, title)
        c.font = white_bold
        c.fill = accent_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    def row(label: str, *vals: str):
        nonlocal r
        ws.cell(r, 1, label).font = bold
        for i, v in enumerate(vals):
            ws.cell(r, 2 + i, v).alignment = wrap
        r += 1

    section("종합 판단")
    row(rec.get("verdict", "—"), rec.get("reason", ""))
    ws.cell(r - 1, 2).alignment = wrap
    r += 1

    section("대지 정보")
    row("용도지역", land.get("zone_use") or "미확인")
    row("지역지구", land.get("zone_district") or "—")
    row("대지면적", f"{_num(result.get('site_area_used'))}㎡ ({_site_source_label(result.get('site_area_source'))})")
    r += 1

    section("이 대지에 지을 수 있는 범위")
    row("최대 건폐율", f"{_num(proposal.get('max_building_coverage_pct'), 1)}%", f"최대 건축면적 {_num(proposal.get('max_building_area_sqm'))}㎡")
    far = proposal.get("far_pct")
    far_relief = proposal.get("max_far_pct_relief")
    row("최대 용적률", f"{_num(far, 1)}%", f"완화 시 최대 {_num(far_relief, 1)}%" if far_relief and far and far_relief > far else "—")
    fa = proposal.get("max_floor_area_sqm")
    fa_relief = proposal.get("max_floor_area_relief_sqm")
    row("가능 연면적", f"{_num(fa)}㎡", f"완화 시 {_num(fa_relief)}㎡" if fa_relief and fa and fa_relief > fa else "용적률 한도 기준")
    row("권장 주차대수", f"{_num(proposal.get('recommended_parking_spaces'))}대", "최대 연면적 기준")
    r += 1

    # 갭 분석
    has_target = any((c.get("gap_analysis") or {}).get("has_target") for c in categories)
    if has_target:
        section("공모 요구치 대비 (갭 분석)")
        for col, txt in enumerate(["카테고리", "공모 요구", "법적 가능"]):
            cell = ws.cell(r, 1 + col, txt)
            cell.font = bold
            cell.fill = header_fill
        r += 1
        for c in categories:
            gap = c.get("gap_analysis") or {}
            if not gap.get("has_target"):
                continue
            unit = c.get("unit", "")
            ws.cell(r, 1, f"{c.get('label')} — {_gap_label(gap.get('status'))}")
            ws.cell(r, 2, f"{_num(c.get('competition_target'), 1)}{unit}")
            ws.cell(r, 3, f"{_num(c.get('legal_limit'), 1)}{unit}")
            r += 1

    # ── Sheet 2: 심의·평가 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("심의평가")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 50
    for col, txt in enumerate(["구분", "항목", "사유"]):
        cell = ws2.cell(1, 1 + col, txt)
        cell.font = white_bold
        cell.fill = accent_fill
    rr = 2
    for kind, items in [("필수", review.get("required") or []), ("조건부", review.get("maybe") or [])]:
        for it in items:
            ws2.cell(rr, 1, kind)
            ws2.cell(rr, 2, it.get("name", "")).alignment = wrap
            ws2.cell(rr, 3, it.get("reason", "")).alignment = wrap
            rr += 1
    if rr == 2:
        ws2.cell(2, 1, "자동 트리거된 심의·평가 없음")

    # ── Sheet 3: 자동 대안 ──────────────────────────────────────────────
    alts = result.get("auto_alternatives") or []
    if alts:
        ws3 = wb.create_sheet("자동대안")
        for col, w in zip("ABCDEF", (16, 10, 10, 14, 12, 40)):
            ws3.column_dimensions[col].width = w
        headers = ["대안", "건폐율", "용적률", "가능연면적", "권장주차", "산정 근거 · 심의"]
        for col, txt in enumerate(headers):
            cell = ws3.cell(1, 1 + col, txt)
            cell.font = white_bold
            cell.fill = accent_fill
        rr3 = 2
        for a in alts:
            d = a.get("derivation") or {}
            notes: list[str] = []
            notes.append(f"기본 {_num(d.get('base_far_pct'), 1)}%")
            for b in d.get("relief_breakdown") or []:
                basis = f" ({b.get('basis')})" if b.get("basis") else ""
                notes.append(f"+{_num(b.get('relief_pct'), 1)}% {b.get('label', '')}{basis}")
            if d.get("cap_note"):
                notes.append(f"캡: {d.get('cap_note')}")
            notes.append(f"→ 최종 {_num(d.get('final_far_pct'), 1)}%")
            reqs = a.get("review_required") or []
            if reqs:
                notes.append(
                    "심의: " + ", ".join(r.get("name", "") for r in reqs)
                )
            ws3.cell(rr3, 1, a.get("label", "")).font = bold
            ws3.cell(rr3, 2, f"{_num(a.get('building_coverage_pct'), 1)}%")
            ws3.cell(rr3, 3, f"{_num(a.get('far_pct'), 1)}%")
            ws3.cell(rr3, 4, f"{_num(a.get('max_floor_area_sqm'))}㎡")
            ws3.cell(rr3, 5, f"{_num(a.get('recommended_parking_spaces'))}대")
            ws3.cell(rr3, 6, "\n".join(notes)).alignment = wrap
            rr3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTML (자체완결 · 인쇄/PDF용) ─────────────────────────────────────────────
def to_html(
    result: dict,
    form_data: dict | None = None,
    project_name: str = "",
    company: str = "",
    author: str = "",
) -> str:
    """사업성 검토 결과 → 자체완결 HTML 보고서.

    LLM 호출 없음. 외부 의존 없는 단일 문서(인쇄·PDF 변환 용이).
    모든 사용자/데이터 문자열은 html.escape 처리.
    """
    import html

    def esc(v: Any) -> str:
        if v is None or v == "":
            return "—"
        return html.escape(str(v))

    today = datetime.now().strftime("%Y-%m-%d")
    address = result.get("address") or "—"
    land = result.get("land_facts") or {}
    proposal = result.get("proposal") or {}
    categories = result.get("categories") or []
    review = result.get("review_burden") or {}
    rec = result.get("overall_recommendation") or {}
    alts = result.get("auto_alternatives") or []

    verdict = rec.get("verdict") or "—"
    verdict_color = {
        "참여 권장": "#16a34a",
        "협상 필요": "#ca8a04",
        "패스 권장": "#dc2626",
        "정보 부족": "#6c757d",
    }.get(verdict, "#6c757d")

    P: list[str] = []
    P.append("<!DOCTYPE html><html lang='ko'><head><meta charset='utf-8'>")
    P.append("<meta name='viewport' content='width=device-width, initial-scale=1'>")
    P.append(f"<title>사전 사업성 검토 — {esc(project_name or address)}</title>")
    P.append("""<style>
:root{--accent:#e60012;--ink:#1f2937;--muted:#6b7280;--line:#e5e7eb;--ok:#16a34a;}
*{box-sizing:border-box;}
body{font-family:'Malgun Gothic','맑은 고딕',-apple-system,sans-serif;color:var(--ink);
  max-width:880px;margin:0 auto;padding:32px 28px;line-height:1.6;font-size:13px;}
h1{font-size:22px;margin:0 0 4px;}
h2{font-size:15px;margin:26px 0 10px;padding-bottom:6px;border-bottom:2px solid var(--accent);}
h3{font-size:13px;margin:16px 0 6px;color:var(--accent);}
.meta{color:var(--muted);font-size:12px;margin-bottom:4px;}
.verdict{display:inline-block;padding:8px 16px;border-radius:8px;font-weight:700;font-size:16px;
  border:2px solid;margin:6px 0;}
table{border-collapse:collapse;width:100%;margin:8px 0;font-size:12px;}
th,td{border:1px solid var(--line);padding:6px 9px;text-align:left;vertical-align:top;}
th{background:#f8f9fa;font-weight:700;}
td.num{text-align:right;font-variant-numeric:tabular-nums;}
.alt{border:1px solid var(--line);border-radius:10px;padding:14px 16px;margin:10px 0;}
.alt-head{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:6px;}
.alt-title{font-size:14px;font-weight:700;}
.tag{font-size:11px;color:var(--muted);}
.derive{background:#f8f9fa;border-radius:8px;padding:10px 12px;margin-top:8px;font-size:12px;}
.derive .step{margin:2px 0;}
.relief{color:var(--ok);font-weight:700;}
.basis{color:var(--muted);}
ul{margin:4px 0;padding-left:18px;}
.foot{margin-top:30px;padding-top:12px;border-top:1px solid var(--line);
  color:var(--muted);font-size:11px;}
@media print{body{padding:0;}h2{page-break-after:avoid;}.alt{page-break-inside:avoid;}}
</style></head><body>""")

    # 헤더
    P.append("<h1>사전 사업성 검토 요약</h1>")
    if project_name:
        P.append(f"<div class='meta'><b>프로젝트</b> {esc(project_name)}</div>")
    P.append(f"<div class='meta'><b>주소</b> {esc(address)}</div>")
    P.append(f"<div class='meta'><b>작성일</b> {esc(today)}")
    if company or author:
        P.append(f" · <b>작성</b> {esc((company + ' ' + author).strip())}")
    P.append("</div>")

    # 종합 판단
    P.append("<h2>종합 판단</h2>")
    P.append(
        f"<div class='verdict' style='color:{verdict_color};border-color:{verdict_color};'>"
        f"{esc(verdict)}</div>"
    )
    P.append(f"<p>{esc(rec.get('reason', ''))}</p>")

    # 대지 정보
    P.append("<h2>대지 정보</h2>")
    P.append("<table>")
    P.append(f"<tr><th>용도지역</th><td>{esc(land.get('zone_use') or '미확인')}</td>"
             f"<th>지역지구</th><td>{esc(land.get('zone_district') or '—')}</td></tr>")
    P.append(
        f"<tr><th>대지면적</th><td class='num'>{_num(result.get('site_area_used'))}㎡</td>"
        f"<th>조회</th><td>{esc(_site_source_label(result.get('site_area_source')))}</td></tr>"
    )
    P.append("</table>")

    # 가능 범위 제안
    P.append("<h2>이 대지에 지을 수 있는 범위</h2>")
    P.append("<table><tr><th>항목</th><th>값</th><th>비고</th></tr>")
    far = proposal.get("far_pct")
    far_relief = proposal.get("max_far_pct_relief")
    fa = proposal.get("max_floor_area_sqm")
    fa_relief = proposal.get("max_floor_area_relief_sqm")
    P.append(
        f"<tr><td>최대 건폐율</td><td class='num'>{_num(proposal.get('max_building_coverage_pct'), 1)}%</td>"
        f"<td>최대 건축면적 {_num(proposal.get('max_building_area_sqm'))}㎡</td></tr>"
    )
    P.append(
        f"<tr><td>최대 용적률</td><td class='num'>{_num(far, 1)}%</td>"
        f"<td>{('완화 시 최대 ' + _num(far_relief, 1) + '%') if far_relief and far and far_relief > far else '—'}</td></tr>"
    )
    P.append(
        f"<tr><td>가능 연면적</td><td class='num'>{_num(fa)}㎡</td>"
        f"<td>{('완화 시 ' + _num(fa_relief) + '㎡') if fa_relief and fa and fa_relief > fa else '용적률 한도 기준'}</td></tr>"
    )
    P.append(
        f"<tr><td>권장 주차대수</td><td class='num'>{_num(proposal.get('recommended_parking_spaces'))}대</td>"
        f"<td>최대 연면적 기준</td></tr>"
    )
    P.append("</table>")

    # 자동 제안 대안 + 산정 근거
    if alts:
        P.append("<h2>자동 제안 대안 (연면적 큰 순)</h2>")
        P.append("<table><tr><th>대안</th><th>건폐율</th><th>용적률</th>"
                 "<th>가능 연면적</th><th>권장 주차</th><th>심의</th></tr>")
        for a in alts:
            P.append(
                f"<tr><td><b>{esc(a.get('label'))}</b></td>"
                f"<td class='num'>{_num(a.get('building_coverage_pct'), 1)}%</td>"
                f"<td class='num'>{_num(a.get('far_pct'), 1)}%</td>"
                f"<td class='num'>{_num(a.get('max_floor_area_sqm'))}㎡</td>"
                f"<td class='num'>{_num(a.get('recommended_parking_spaces'))}대</td>"
                f"<td class='num'>{a.get('review_count_required', 0)}건</td></tr>"
            )
        P.append("</table>")

        for a in alts:
            d = a.get("derivation") or {}
            P.append("<div class='alt'>")
            P.append(
                f"<div class='alt-head'><span class='alt-title'>{esc(a.get('label'))}</span>"
                f"<span class='tag'>{esc(a.get('tagline'))}</span></div>"
            )
            P.append("<div class='derive'>")
            P.append("<div><b>용적률 산정</b></div>")
            P.append(f"<div class='step'>기본 한도 <b>{_num(d.get('base_far_pct'), 1)}%</b>"
                     f"{(' (' + esc(d.get('far_source')) + ')') if d.get('far_source') else ''}</div>")
            bd = d.get("relief_breakdown") or []
            if bd:
                P.append("<ul>")
                for b in bd:
                    basis = f" <span class='basis'>· {esc(b.get('basis'))}</span>" if b.get("basis") else ""
                    note = f" <span class='basis'>· {esc(b.get('note'))}</span>" if b.get("note") else ""
                    P.append(
                        f"<li><span class='relief'>+{_num(b.get('relief_pct'), 1)}%</span> "
                        f"{esc(b.get('label'))}{basis}{note}</li>"
                    )
                P.append("</ul>")
            else:
                P.append("<div class='step basis'>완화 적용 없음 (기본 한도)</div>")
            if d.get("cap_note"):
                P.append(f"<div class='step'>⚖ {esc(d.get('cap_note'))}</div>")
            P.append(f"<div class='step'><b>→ 최종 용적률 {_num(d.get('final_far_pct'), 1)}%</b></div>")
            if d.get("site_area_sqm"):
                P.append(
                    f"<div class='step'>연면적 = {_num(d.get('site_area_sqm'))}㎡ × "
                    f"{_num(d.get('final_far_pct'), 1)}% = <b>{_num(a.get('max_floor_area_sqm'))}㎡</b></div>"
                )
            reqs = a.get("review_required") or []
            if reqs:
                P.append(f"<div class='step' style='margin-top:6px;'><b>심의·평가 필요 ({len(reqs)})</b></div><ul>")
                for r in reqs:
                    ref = f" <span class='basis'>「{esc(r.get('law_ref'))}」</span>" if r.get("law_ref") else ""
                    reason = f" — {esc(r.get('reason'))}" if r.get("reason") else ""
                    P.append(f"<li>{esc(r.get('name'))}{reason}{ref}</li>")
                P.append("</ul>")
            P.append("</div></div>")

    # 갭 분석
    has_target = any((c.get("gap_analysis") or {}).get("has_target") for c in categories)
    if has_target:
        P.append("<h2>공모 요구치 대비 (갭 분석)</h2>")
        P.append("<table><tr><th>카테고리</th><th>공모 요구</th><th>법적 가능</th><th>판정</th></tr>")
        for c in categories:
            gap = c.get("gap_analysis") or {}
            if not gap.get("has_target"):
                continue
            unit = c.get("unit", "")
            P.append(
                f"<tr><td>{esc(c.get('label'))}</td>"
                f"<td class='num'>{_num(c.get('competition_target'), 1)}{esc(unit)}</td>"
                f"<td class='num'>{_num(c.get('legal_limit'), 1)}{esc(unit)}</td>"
                f"<td>{esc(_gap_label(gap.get('status')))} ({esc(gap.get('gap_text', ''))})</td></tr>"
            )
        P.append("</table>")

    # 심의 부담
    P.append("<h2>심의·평가 부담</h2>")
    req_items = review.get("required") or []
    maybe_items = review.get("maybe") or []
    if req_items or maybe_items:
        if req_items:
            P.append(f"<h3>필수 ({len(req_items)})</h3><ul>")
            for r in req_items:
                ref = f" <span class='basis'>「{esc(r.get('law_ref'))}」</span>" if r.get("law_ref") else ""
                P.append(f"<li>{esc(r.get('name'))} — {esc(r.get('reason'))}{ref}</li>")
            P.append("</ul>")
        if maybe_items:
            P.append(f"<h3>조건부 ({len(maybe_items)})</h3><ul>")
            for r in maybe_items:
                P.append(f"<li>{esc(r.get('name'))} — {esc(r.get('reason'))}</li>")
            P.append("</ul>")
    else:
        P.append("<p>자동 트리거된 심의·평가 없음.</p>")

    P.append(
        "<div class='foot'>본 사전 사업성 검토는 시행령·조례·고시 원문 기준 자동 산정값입니다. "
        "참여 판단 보조용이며, 실제 인허가 한도·가능성은 도시계획·건축 심의 및 "
        "시니어 건축사 검토로 확정됩니다.</div>"
    )
    P.append("</body></html>")
    return "".join(P)
