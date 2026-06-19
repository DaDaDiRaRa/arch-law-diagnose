"""진단 결과 → Markdown / Excel 내보내기.

LLM 호출 없음. 진단 결과(dict)를 받아 텍스트·파일로 변환만 한다.
LegalReviewReport HTML 구조와 같은 정보를 다른 포맷으로 표현.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)


_CATEGORY_LABELS = {
    "행위제한": "행위제한 적합성",
    "도시계획시설": "도시계획시설 저촉",
    "건폐율": "건폐율",
    "용적률": "용적률",
    "높이_일조": "높이·일조",
    "주차": "주차",
    "조경": "조경",
    "설비_소방": "설비·소방",
    "공공시설_의무인증": "공공시설 의무 인증",
    "BF_인증": "BF 인증 (무장애)",
    "범죄예방_건축기준": "범죄예방 건축기준",
    "다중이용건축물": "다중이용건축물 분류",
    "중첩지구_구역": "중첩 지구·구역",
    "철도보호지구": "철도보호지구",
}


def _pass_label(p: Any) -> str:
    if p is True:
        return "적법함"
    if p is False:
        return "부적법"
    return "확인필요"


def _signal_label(signal: str) -> str:
    return {"GREEN": "🟢 적합", "YELLOW": "🟡 주의", "RED": "🔴 부적합"}.get(
        signal or "", signal or "—"
    )


def _g(obj: Any, key: str, default: Any = None) -> Any:
    """dict 가정 .get() 안전 호출 — dict가 아니면 default 반환."""
    if isinstance(obj, dict):
        return obj.get(key, default)
    return default


def _review_items(applicable_reviews: Any) -> list[dict]:
    """applicable_reviews는 dict({items: [...]}) 또는 list 둘 다 허용.

    review_triggers.evaluate_reviews()의 실제 반환은 dict 형태.
    """
    if isinstance(applicable_reviews, dict):
        items = applicable_reviews.get("items")
        if isinstance(items, list):
            return [it for it in items if isinstance(it, dict)]
    elif isinstance(applicable_reviews, list):
        return [it for it in applicable_reviews if isinstance(it, dict)]
    return []


def _review_reason(r: dict) -> str:
    """review 항목의 사유 — triggered_reasons (list[str])를 우선 사용."""
    reasons = r.get("triggered_reasons")
    if isinstance(reasons, list) and reasons:
        return " · ".join(str(x) for x in reasons if x)
    return str(r.get("reason") or r.get("note") or "—")


def _review_status(r: dict) -> str:
    """review 항목의 판정 — review_triggers는 'severity' 키 사용."""
    s = _g(r, "severity") or _g(r, "status") or ""
    return str(s).upper()


def _unwrap_multi(raw_result: dict) -> tuple[dict, dict | None]:
    """multi_parcel 모드면 (result, multi_info) 분리."""
    if raw_result.get("mode") == "multi_parcel":
        return raw_result.get("result", {}), {
            "parcels": raw_result.get("parcels", []),
            "aggregate": raw_result.get("aggregate", {}),
        }
    return raw_result, None


# ── Markdown 내보내기 ──────────────────────────────────────────────────────


def to_markdown(
    raw_result: dict,
    form_data: dict | None = None,
    project_name: str = "",
    company: str = "",
    author: str = "",
) -> str:
    """진단 결과 → Markdown 텍스트."""
    result, multi = _unwrap_multi(raw_result)
    today = datetime.now().strftime("%Y-%m-%d")

    lines: list[str] = []

    # ── 헤더 ───────────────────────────────────────────────────────────────
    lines.append("# 법규 검토서")
    lines.append("")
    lines.append(f"| 프로젝트명 | {project_name or '—'} | 작성일 | {today} |")
    lines.append("| --- | --- | --- | --- |")
    addr = result.get("address") or (form_data or {}).get("address") or "—"
    lines.append(f"| 대지 주소 | {addr} | | |")
    bu = (form_data or {}).get("building_use") or "—"
    bu_detail = (form_data or {}).get("building_use_detail")
    if bu_detail:
        bu = f"{bu} · {bu_detail}"
    lines.append(f"| 건축물 용도 | {bu} | | |")
    land = result.get("land_info") or {}
    zone_use = land.get("zone_use") or "—"
    zone_district = land.get("zone_district") or "—"
    lines.append(f"| 용도지역 | {zone_use} | 지역지구 | {zone_district} |")

    site = result.get("site_correction") or {}
    if site.get("applied"):
        eff = site.get("effective_m2", 0)
        orig = site.get("original_m2", 0)
        excl = site.get("excluded_m2", 0)
        size_str = f"{eff:,.1f}㎡ (입력 {orig:,.0f}㎡ - 시설부지 {excl:,.1f}㎡, 시행령 §3)"
    else:
        size_str = f"{(form_data or {}).get('site_area') or '—'}㎡"
    lines.append(
        f"| 대지면적 | {size_str} | 종합 판정 | {_signal_label(result.get('signal'))} · "
        f"{result.get('overall_score', '—')}/10 |"
    )
    lines.append("")

    # ── 대지면적 보정 ─────────────────────────────────────────────────────
    if site.get("applied"):
        lines.append("## 대지면적 보정 내역 (도시계획시설 저촉)")
        lines.append("")
        src = site.get("source") or "자동"
        lines.append(f"산정 근거: **{src}** — {site.get('note', '')}")
        lines.append("")
        by_fac = site.get("by_facility") or []
        by_fac = [f for f in by_fac if isinstance(f, dict)]
        if by_fac:
            lines.append("| 구분 | 시설명 | 저촉 면적(㎡) |")
            lines.append("| --- | --- | ---: |")
            for f in by_fac:
                cat = _g(f, "category") or _g(f, "facility_class") or "—"
                name = _g(f, "facility_name") or _g(f, "name") or "—"
                area = _g(f, "area_m2", _g(f, "overlap_area_m2", 0)) or 0
                lines.append(f"| {cat} | {name} | {float(area):,.2f} |")
            lines.append("")

    # ── 인허가 심의 트리거 ───────────────────────────────────────────────
    reviews = _review_items(result.get("applicable_reviews"))
    if reviews:
        required = [r for r in reviews if _review_status(r) == "REQUIRED"]
        maybe = [r for r in reviews if _review_status(r) == "MAYBE"]
        lines.append(
            f"## 인허가 심의 트리거 (필요 {len(required)}건 · 검토 {len(maybe)}건)"
        )
        lines.append("")
        lines.append("| 심의명 | 판정 | 트리거 사유 | 근거법령 |")
        lines.append("| --- | --- | --- | --- |")
        for r in reviews:
            status = _review_status(r)
            badge = (
                "**필요**" if status == "REQUIRED"
                else "검토" if status == "MAYBE"
                else "해당없음"
            )
            name = _g(r, "name") or _g(r, "category") or "—"
            reason = _review_reason(r).replace("|", "\\|")
            law = (_g(r, "law_ref") or _g(r, "basis") or "—").replace("|", "\\|")
            lines.append(f"| {name} | {badge} | {reason} | {law} |")
        lines.append("")

    # ── 건축 개요 ─────────────────────────────────────────────────────────
    if form_data:
        lines.append("## 건축 개요")
        lines.append("")
        fa_above = float(form_data.get("floor_area_above") or 0)
        fa_below = float(form_data.get("floor_area_below") or 0)
        ba = float(form_data.get("building_area") or 0)
        fa_above_f = f"{fa_above:,.2f}㎡" if fa_above else "—"
        fa_below_f = f"{fa_below:,.2f}㎡" if fa_below else "—"
        ba_f = f"{ba:,.2f}㎡" if ba else "—"
        total = fa_above + fa_below
        lines.append(f"| 건축면적 | {ba_f} | 연면적 (지상+지하) | {total:,.2f}㎡ |")
        lines.append("| --- | --- | --- | --- |")
        lines.append(f"| 지상 연면적 | {fa_above_f} | 지하 연면적 | {fa_below_f} |")
        lines.append(
            f"| 층수 / 높이 | 지상 {form_data.get('floors_above') or '—'}층 / "
            f"지하 {form_data.get('floors_below') or '0'}층 / "
            f"높이 {form_data.get('height') or '—'}m | | |"
        )
        lines.append("")

    # ── 법규 검토 결과 ───────────────────────────────────────────────────
    lines.append("## 법규 검토 결과")
    lines.append("")
    lines.append("| 항목 | 적법 여부 | 점수 | 비고 |")
    lines.append("| --- | --- | --- | --- |")
    for cat, data in (result.get("results") or {}).items():
        label = _CATEGORY_LABELS.get(cat, cat)
        passed = _pass_label(data.get("pass"))
        score = data.get("score")
        score_str = f"{score}/10" if score is not None else "—"
        notes = (data.get("notes") or "").replace("\n", " ").replace("|", "\\|")
        if len(notes) > 200:
            notes = notes[:200] + "…"
        lines.append(f"| **{label}** | {passed} | {score_str} | {notes} |")
    lines.append("")

    # ── 검토 의견 (위험 + 경고) ─────────────────────────────────────────
    risks = [r for r in (result.get("risks") or []) if isinstance(r, dict)]
    warnings = [w for w in (result.get("warnings") or []) if isinstance(w, dict)]
    if risks:
        lines.append(f"## 위험 항목 ({len(risks)}건)")
        lines.append("")
        for r in risks:
            cat = _g(r, "category", "")
            cat_label = _CATEGORY_LABELS.get(cat, cat)
            reason = str(_g(r, "reason", "")).strip()
            lines.append(f"- **{cat_label}**: {reason}")
        lines.append("")
    if warnings:
        lines.append(f"## 검토 필요 ({len(warnings)}건)")
        lines.append("")
        for w in warnings:
            cat = _g(w, "category", "")
            cat_label = _CATEGORY_LABELS.get(cat, cat)
            reason = str(_g(w, "reason", "")).strip()
            lines.append(f"- **{cat_label}**: {reason}")
        lines.append("")

    # ── 데이터 품질 ───────────────────────────────────────────────────────
    dq = result.get("data_quality") or {}
    issues = dq.get("issues") or []
    if issues or dq:
        lines.append("## 데이터 품질·출처")
        lines.append("")
        sources = []
        sources.append(f"용도지역: {dq.get('zone_use_source', '—')}")
        sources.append(f"조례: {'적용' if dq.get('ordinance_used') else '미적용'}")
        sources.append(f"LURIS: {'활성' if dq.get('luris_used') else '비활성'}")
        sources.append(f"AI(Claude): {'활성' if dq.get('llm_used') else '비활성'}")
        lines.append(" · ".join(sources))
        lines.append("")
        if issues:
            lines.append("| 구분 | 주의 사항 |")
            lines.append("| --- | --- |")
            for iss in issues:
                level = _g(iss, "level", "info")
                msg = str(_g(iss, "msg", "")).replace("|", "\\|")
                lines.append(f"| {level} | {msg} |")
            lines.append("")

    # ── 합필 정보 (multi_parcel 모드일 때) ───────────────────────────────
    if multi:
        agg = multi.get("aggregate") or {}
        parcels = multi.get("parcels") or []
        lines.append("## 합필 정보")
        lines.append("")
        lines.append(
            f"- 합산 모드: {agg.get('calc_mode', '—')} · "
            f"총 면적: {agg.get('total_site_area', 0):,.1f}㎡ · "
            f"기준 용도지역: {agg.get('primary_zone', '—')}"
        )
        if parcels:
            lines.append("")
            lines.append("| 주소 | PNU | 면적(㎡) | 용도지역 | 관할 |")
            lines.append("| --- | --- | ---: | --- | --- |")
            for p in parcels:
                if not isinstance(p, dict):
                    continue
                lines.append(
                    f"| {_g(p, 'address', '—')} | {_g(p, 'pnu', '—')} | "
                    f"{float(_g(p, 'site_area', 0) or 0):,.0f} | {_g(p, 'zone_use', '—')} | "
                    f"{_g(p, 'jurisdiction_name', '—')} |"
                )
            lines.append("")

    # ── 푸터 ──────────────────────────────────────────────────────────────
    if company or author:
        lines.append(f"---")
        lines.append("")
        lines.append(f"회사: {company or '—'} · 작성자: {author or '—'}")
        lines.append("")
    lines.append(
        "> 본 검토서는 arch-law-diagnose 자동 진단 시스템에 의해 작성되었으며, "
        "최종 법규 해석은 반드시 시니어 검토자/설계자가 확인해야 합니다."
    )
    lines.append(f"> 자동 진단 일자: {today}")

    return "\n".join(lines) + "\n"


# ── Excel 내보내기 ────────────────────────────────────────────────────────


def to_xlsx(
    raw_result: dict,
    form_data: dict | None = None,
    project_name: str = "",
    company: str = "",
    author: str = "",
) -> bytes:
    """진단 결과 → xlsx bytes. 4개 시트 구성."""
    # openpyxl은 무거우니 lazy import
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    result, multi = _unwrap_multi(raw_result)
    today = datetime.now().strftime("%Y-%m-%d")

    wb = Workbook()

    # 공통 스타일
    header_fill = PatternFill(start_color="F1F3F5", end_color="F1F3F5", fill_type="solid")
    accent_fill = PatternFill(start_color="E60012", end_color="E60012", fill_type="solid")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Sheet 1: 종합 ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "종합"
    ws["A1"] = "법규 검토서"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = center

    row = 3
    rows = [
        ("프로젝트명", project_name or "—", "작성일", today),
        ("대지 주소", result.get("address") or (form_data or {}).get("address") or "—", "", ""),
        ("건축물 용도", (form_data or {}).get("building_use") or "—", "신청 주체", (form_data or {}).get("applicant_type") or "—"),
    ]
    land = result.get("land_info") or {}
    rows.append(("용도지역", land.get("zone_use") or "—", "지역지구", land.get("zone_district") or "—"))
    site = result.get("site_correction") or {}
    if site.get("applied"):
        size_text = (
            f"{site.get('effective_m2', 0):,.1f}㎡ "
            f"(입력 {site.get('original_m2', 0):,.0f} - 시설부지 {site.get('excluded_m2', 0):,.1f})"
        )
    else:
        size_text = f"{(form_data or {}).get('site_area') or '—'}㎡"
    signal_text = (
        f"{_signal_label(result.get('signal'))} · {result.get('overall_score', '—')}/10"
    )
    rows.append(("대지면적", size_text, "종합 판정", signal_text))

    for label, val, label2, val2 in rows:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value=val).alignment = wrap
        if label2:
            ws.cell(row=row, column=3, value=label2).font = bold
            ws.cell(row=row, column=3).fill = header_fill
            ws.cell(row=row, column=4, value=val2).alignment = wrap
        else:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="검토 의견").font = Font(bold=True, size=13)
    row += 1
    risks = [r for r in (result.get("risks") or []) if isinstance(r, dict)]
    warnings = [w for w in (result.get("warnings") or []) if isinstance(w, dict)]
    if risks:
        ws.cell(row=row, column=1, value=f"위험 항목 ({len(risks)}건)").font = Font(bold=True, color="DC2626")
        row += 1
        for r in risks:
            cat = _g(r, "category", "")
            ws.cell(row=row, column=1, value=_CATEGORY_LABELS.get(cat, cat))
            ws.cell(row=row, column=2, value=str(_g(r, "reason", ""))).alignment = wrap
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1
    if warnings:
        row += 1
        ws.cell(row=row, column=1, value=f"검토 필요 ({len(warnings)}건)").font = Font(bold=True, color="CA8A04")
        row += 1
        for w in warnings:
            cat = _g(w, "category", "")
            ws.cell(row=row, column=1, value=_CATEGORY_LABELS.get(cat, cat))
            ws.cell(row=row, column=2, value=str(_g(w, "reason", ""))).alignment = wrap
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

    for col_idx, width in enumerate([18, 40, 18, 40], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 2: 법규 검토 결과 ──────────────────────────────────────────
    ws2 = wb.create_sheet("법규 검토")
    headers = ["항목", "적법 여부", "점수", "신뢰도", "출처", "비고"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = accent_fill
        cell.font = white_bold
        cell.alignment = center
    ws2.freeze_panes = "A2"

    row = 2
    for cat, data in (result.get("results") or {}).items():
        label = _CATEGORY_LABELS.get(cat, cat)
        score = _g(data, "score")
        score_str = f"{score}/10" if score is not None else "—"
        ws2.cell(row=row, column=1, value=label).font = bold
        ws2.cell(row=row, column=2, value=_pass_label(_g(data, "pass")))
        ws2.cell(row=row, column=3, value=score_str)
        ws2.cell(row=row, column=4, value=_g(data, "confidence", "—"))
        ws2.cell(row=row, column=5, value=_g(data, "source") or "—").alignment = wrap
        ws2.cell(row=row, column=6, value=_g(data, "notes") or "").alignment = wrap
        ws2.row_dimensions[row].height = 60
        row += 1

    for col_idx, width in enumerate([18, 12, 8, 8, 28, 60], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 3: 심의 트리거 ─────────────────────────────────────────────
    ws3 = wb.create_sheet("심의 트리거")
    headers = ["심의명", "판정", "트리거 사유", "근거법령"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.fill = accent_fill
        cell.font = white_bold
        cell.alignment = center
    ws3.freeze_panes = "A2"

    row = 2
    reviews = _review_items(result.get("applicable_reviews"))
    for r in reviews:
        status = _review_status(r)
        badge = "필요" if status == "REQUIRED" else "검토" if status == "MAYBE" else "해당없음"
        ws3.cell(row=row, column=1, value=_g(r, "name") or _g(r, "category") or "—").font = bold
        ws3.cell(row=row, column=2, value=badge)
        ws3.cell(row=row, column=3, value=_review_reason(r)).alignment = wrap
        ws3.cell(row=row, column=4, value=_g(r, "law_ref") or _g(r, "basis") or "—").alignment = wrap
        ws3.row_dimensions[row].height = 50
        row += 1

    for col_idx, width in enumerate([22, 10, 50, 30], start=1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 4: 데이터 품질 ─────────────────────────────────────────────
    ws4 = wb.create_sheet("데이터 품질")
    dq = result.get("data_quality") or {}
    ws4["A1"] = "데이터 출처·품질"
    ws4["A1"].font = Font(bold=True, size=13)
    ws4.merge_cells("A1:C1")

    row = 3
    src_rows = [
        ("용도지역 출처", dq.get("zone_use_source", "—")),
        ("조례 적용", "예" if dq.get("ordinance_used") else "아니오 — 시행령 기본값"),
        ("LURIS API", "활성" if dq.get("luris_used") else "비활성"),
        ("AI (Claude)", "활성" if dq.get("llm_used") else "비활성"),
        ("도로폭 출처", dq.get("road_width_source", "—")),
        ("도로폭 값", dq.get("road_width_used", "—")),
        ("토지 캐시 stale", "예" if dq.get("land_cache_stale") else "아니오"),
        ("캐시 경과 일수", dq.get("land_cache_age_days", 0)),
    ]
    for label, val in src_rows:
        ws4.cell(row=row, column=1, value=label).font = bold
        ws4.cell(row=row, column=1).fill = header_fill
        ws4.cell(row=row, column=2, value=str(val))
        row += 1

    issues = dq.get("issues") or []
    if issues:
        row += 1
        ws4.cell(row=row, column=1, value=f"품질 알림 ({len(issues)}건)").font = Font(bold=True, size=12)
        row += 1
        for h_idx, h in enumerate(["구분", "코드", "내용"], start=1):
            c = ws4.cell(row=row, column=h_idx, value=h)
            c.fill = header_fill
            c.font = bold
        row += 1
        for iss in issues:
            ws4.cell(row=row, column=1, value=_g(iss, "level", "info"))
            ws4.cell(row=row, column=2, value=_g(iss, "code", ""))
            ws4.cell(row=row, column=3, value=str(_g(iss, "msg", ""))).alignment = wrap
            row += 1

    for col_idx, width in enumerate([20, 18, 60], start=1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 출력 ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
